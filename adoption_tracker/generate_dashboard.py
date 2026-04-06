#!/usr/bin/env python3
"""
generate_dashboard.py
=====================
Reads wms_adoption_results.xlsx and produces a self-contained index.html
suitable for GitHub Pages deployment.

Usage:
    python3 generate_dashboard.py
    python3 generate_dashboard.py --input path/to/results.xlsx --output docs/index.html
"""

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("❌  openpyxl not installed — run: pip install openpyxl")

# Import feature descriptions from the single source of truth
sys.path.insert(0, str(Path(__file__).parent))
try:
    from features import FEATURES as _FEATURES
    FEATURE_DESCRIPTIONS: dict = {
        f["feature"]: f["description"]
        for f in _FEATURES
        if "description" in f
    }
except Exception:
    FEATURE_DESCRIPTIONS = {}  # graceful fallback if features.py unavailable

HERE           = Path(__file__).parent
DEFAULT_INPUT  = HERE / "output" / "wms_adoption_results.xlsx"
DEFAULT_OUTPUT = HERE.parent / "docs" / "index.html"


# ─── DATA LOADING ─────────────────────────────────────────────────────────────

def load_results(xlsx_path: Path) -> list:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Detail" not in wb.sheetnames:
        sys.exit("❌  'Detail' sheet not found in workbook.")
    ws      = wb["Detail"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows    = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or str(row[0]).strip().upper() == "TOTAL":
            continue                        # skip blank rows and the Excel summary row
        d = dict(zip(headers, row))
        rows.append({
            "market":       str(d.get("Market", "")),
            "warehouse":    str(d.get("Warehouse", "")),
            "feature":      str(d.get("Feature", "")),
            "domain":       str(d.get("Domain", "")),
            "signal":       str(d.get("Signal", "")),
            "signal_value": str(d.get("Signal Value", "")),
            "adopted":      str(d.get("Adopted", "Error")),  # "Yes"/"No"/"Error"
            "checked_at":   str(d.get("Checked At", "")),
        })
    return rows


# ─── DATA TRANSFORMATION ──────────────────────────────────────────────────────

# Domain display order — anything not listed falls to the end alphabetically
DOMAIN_ORDER = ["Outbound", "Inbound", "Inventory", "Move Transfer", "RMS", "Multi-WH"]


def _domain_rank(domain: str) -> int:
    try:
        return DOMAIN_ORDER.index(domain)
    except ValueError:
        return len(DOMAIN_ORDER)


def build_payload(results: list) -> dict:
    markets   = sorted(set(r["market"]   for r in results))

    # Build domain map and filter out features with no domain (empty / "None")
    domains_raw = {r["feature"]: r["domain"] for r in results}
    valid_domains = {f: d for f, d in domains_raw.items()
                     if d and d.strip().lower() not in ("", "none")}

    # Stable insertion-ordered feature list, filtered, then sorted by domain rank
    seen_features = list(dict.fromkeys(r["feature"] for r in results))
    features = sorted(
        [f for f in seen_features if f in valid_domains],
        key=lambda f: (_domain_rank(valid_domains[f]), f)
    )

    wh_by_mkt = defaultdict(set)
    for r in results:
        wh_by_mkt[r["market"]].add(r["warehouse"])
    wh_by_mkt = {m: sorted(whs) for m, whs in sorted(wh_by_mkt.items())}

    # Feature × Market summary  →  { feat: { mkt: {yes,total,pct} | None } }
    feat_mkt: dict = {}
    for feat in features:
        feat_mkt[feat] = {}
        for mkt in markets:
            rows = [r for r in results if r["feature"] == feat and r["market"] == mkt]
            if not rows:
                feat_mkt[feat][mkt] = None
            else:
                yes   = sum(1 for r in rows if r["adopted"] == "Yes")
                total = len(rows)
                feat_mkt[feat][mkt] = {
                    "yes": yes, "total": total,
                    "pct": round(yes / total * 100) if total else 0,
                }

    # Feature-level total (aggregated across ALL markets)
    feat_total: dict = {}
    for feat in features:
        rows = [r for r in results if r["feature"] == feat]
        yes   = sum(1 for r in rows if r["adopted"] == "Yes")
        total = len(rows)
        feat_total[feat] = {
            "yes": yes, "total": total,
            "pct": round(yes / total * 100) if total else 0,
        }

    # Warehouse × Feature cells  →  { "mkt|wh|feat": "Yes"/"No"/"Error" }
    cells: dict = {}
    for r in results:
        if r["feature"] not in valid_domains:
            continue
        key = f"{r['market']}|{r['warehouse']}|{r['feature']}"
        cells[key] = r["adopted"]

    # Overall stats (excluding filtered-out features)
    filtered = [r for r in results if r["feature"] in valid_domains]
    total_f   = len(filtered)
    adopted_f = sum(1 for r in filtered if r["adopted"] == "Yes")
    errors_f  = sum(1 for r in filtered if r["adopted"] == "Error")
    checked_at = results[0]["checked_at"][:10] if results else "—"

    # Adoption-definition tooltips (fall back to generic message if not defined)
    feat_desc: dict = {
        f: FEATURE_DESCRIPTIONS.get(
            f, "Adoption definition not yet documented for this feature."
        )
        for f in features
    }

    return {
        "markets":     markets,
        "features":    features,
        "domains":     valid_domains,
        "feat_desc":   feat_desc,
        "wh_by_mkt":   wh_by_mkt,
        "feat_mkt":    feat_mkt,
        "feat_total":  feat_total,
        "cells":       cells,
        "stats": {
            "total":       total_f,
            "adopted":     adopted_f,
            "not_adopted": total_f - adopted_f - errors_f,
            "errors":      errors_f,
            "pct":         round(adopted_f / total_f * 100) if total_f else 0,
            "wh_count":    sum(len(v) for v in wh_by_mkt.values()),
            "mkt_count":   len(markets),
            "feat_count":  len(features),
            "checked_at":  checked_at,
        },
    }


# ─── HTML GENERATION ──────────────────────────────────────────────────────────

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>WMS Feature Adoption Dashboard</title>
<script src="https://cdn.tailwindcss.com"></script>
<style>
  body { font-family: 'Inter', system-ui, sans-serif; }
  .cell-yes   { background:#dcfce7; color:#166534; font-weight:600; }
  .cell-no    { background:#fee2e2; color:#991b1b; font-weight:600; }
  .cell-err   { background:#fef9c3; color:#854d0e; font-weight:600; }
  .cell-na    { background:#f3f4f6; color:#9ca3af; }
  .pct-100    { background:#15803d; color:#fff; }
  .pct-high   { background:#22c55e; color:#fff; }
  .pct-mid    { background:#f59e0b; color:#fff; }
  .pct-low    { background:#ef4444; color:#fff; }
  .pct-zero   { background:#dc2626; color:#fff; }
  .pct-na     { background:#e5e7eb; color:#9ca3af; }
  .tab-active { background:#ee4d2d; color:#fff; }
  .tab-idle   { background:#f3f4f6; color:#6b7280; cursor:pointer; }
  .tab-idle:hover { background:#e5e7eb; }
  table { border-collapse:collapse; }
  th, td { border:1px solid #e5e7eb; }
  thead th { background:#f9fafb; position:sticky; top:0; z-index:2; }
  .sticky-col { position:sticky; left:0; z-index:1; background:#fff; }
  thead .sticky-col { z-index:3; }
  .scroll-x { overflow-x:auto; border:1px solid #e5e7eb; border-radius:8px; }

  /* ── Tooltip icon ── */
  .tip-wrap { display:inline-flex; align-items:center; gap:4px; cursor:default; }
  .tip-icon { display:inline-flex; align-items:center; justify-content:center;
              width:14px; height:14px; border-radius:50%; background:#94a3b8;
              color:#fff; font-size:9px; font-weight:700; flex-shrink:0;
              cursor:help; line-height:1; }
  /* ── Floating tooltip portal (appended to <body> via JS) ── */
  #tip-portal { display:none; position:fixed; z-index:9999;
                background:#1e293b; color:#e2e8f0; font-size:11px; line-height:1.6;
                padding:9px 12px; border-radius:7px; max-width:280px;
                box-shadow:0 6px 20px rgba(0,0,0,.4); font-weight:400;
                white-space:normal; pointer-events:none; }
</style>
</head>
<body class="bg-gray-50 min-h-screen">

<!-- ── Floating tooltip portal (never clipped by table cells) ─── -->
<div id="tip-portal"></div>

<!-- ── Header ─────────────────────────────────────────────────── -->
<div class="bg-white border-b border-gray-200 shadow-sm">
  <div class="max-w-screen-xl mx-auto px-6 py-4 flex items-center justify-between flex-wrap gap-3">
    <div class="flex items-center gap-3">
      <span class="text-2xl">🏭</span>
      <div>
        <h1 class="text-xl font-bold text-gray-900">WMS Feature Adoption</h1>
        <p class="text-xs text-gray-400">Last checked: <span id="checkedAt"></span></p>
      </div>
    </div>
    <div class="flex gap-4 flex-wrap" id="kpiRow"></div>
  </div>
</div>

<!-- ── Info Banner ────────────────────────────────────────────── -->
<div class="max-w-screen-xl mx-auto px-6 mt-5">
  <div class="grid grid-cols-1 md:grid-cols-3 gap-3">

    <div class="flex gap-3 items-start bg-white border border-blue-100 rounded-xl px-4 py-3 shadow-sm">
      <span class="text-xl mt-0.5">📅</span>
      <div>
        <p class="text-xs font-semibold text-blue-700 uppercase tracking-wide mb-0.5">Data Update</p>
        <p class="text-sm text-gray-700">
          Updated <strong>weekly</strong>. Last update:
          <span class="font-semibold text-gray-900" id="bannerDate"></span>
        </p>
      </div>
    </div>

    <div class="flex gap-3 items-start bg-white border border-green-100 rounded-xl px-4 py-3 shadow-sm">
      <span class="text-xl mt-0.5">🏢</span>
      <div>
        <p class="text-xs font-semibold text-green-700 uppercase tracking-wide mb-0.5">Warehouse Scope</p>
        <p class="text-sm text-gray-700">
          Active warehouses only — those with <strong>≥ 1 inbound or outbound order</strong>
          in the past 7 days.
        </p>
      </div>
    </div>

    <div class="flex gap-3 items-start bg-white border border-orange-100 rounded-xl px-4 py-3 shadow-sm">
      <span class="text-xl mt-0.5">💡</span>
      <div>
        <p class="text-xs font-semibold text-orange-700 uppercase tracking-wide mb-0.5">Track a New Feature?</p>
        <p class="text-sm text-gray-700">
          Adding a feature requires testing before it goes live.
          Contact <strong>zhengyu.chenzy</strong> to request tracking.
        </p>
      </div>
    </div>

  </div>
</div>

<!-- ── Tabs ───────────────────────────────────────────────────── -->
<div class="max-w-screen-xl mx-auto px-6 mt-6">
  <div class="flex gap-2 flex-wrap" id="mainTabs">
    <button onclick="showSection('overview')"  class="tab-active  px-4 py-2 rounded-lg text-sm font-medium" id="tab-overview">📊 Overview</button>
    <button onclick="showSection('warehouse')" class="tab-idle px-4 py-2 rounded-lg text-sm font-medium" id="tab-warehouse">🏢 Warehouse Detail</button>
  </div>
</div>

<!-- ── Overview Section ───────────────────────────────────────── -->
<div id="sec-overview" class="max-w-screen-xl mx-auto px-6 mt-4">
  <p class="text-sm text-gray-500 mb-3">Adoption rate per feature per market. Green ≥ 80% · Amber 40–79% · Red &lt; 40%</p>
  <div class="scroll-x">
    <table class="w-full text-sm" id="overviewTable"></table>
  </div>
</div>

<!-- ── Warehouse Detail Section ───────────────────────────────── -->
<div id="sec-warehouse" class="max-w-screen-xl mx-auto px-6 mt-4 hidden">
  <div class="flex gap-2 flex-wrap mb-4" id="mktTabs"></div>
  <div id="wh-tables"></div>
</div>

<!-- ── Legend ─────────────────────────────────────────────────── -->
<div class="max-w-screen-xl mx-auto px-6 mt-6 pb-10">
  <div class="flex flex-wrap gap-4 text-xs text-gray-500 items-center">
    <span class="font-semibold text-gray-600">Legend:</span>
    <span class="flex items-center gap-1"><span class="inline-block w-3 h-3 rounded-sm bg-green-600"></span> Adopted (≥80%)</span>
    <span class="flex items-center gap-1"><span class="inline-block w-3 h-3 rounded-sm bg-yellow-400"></span> Partial (40–79%)</span>
    <span class="flex items-center gap-1"><span class="inline-block w-3 h-3 rounded-sm bg-red-500"></span> Low / None (&lt;40%)</span>
    <span class="flex items-center gap-1"><span class="inline-block w-3 h-3 rounded-sm bg-gray-200"></span> Not applicable</span>
    <span class="ml-4">Warehouse cells: ✓ Adopted · ✗ Not adopted · ! Error</span>
  </div>
</div>

<script>
const DATA = __DATA__;

/* ── KPI row + banner date ────────────────────────────────────── */
const s = DATA.stats;
document.getElementById("checkedAt").textContent = s.checked_at;
document.getElementById("bannerDate").textContent = s.checked_at;
document.getElementById("kpiRow").innerHTML = [
  ["Markets",    s.mkt_count,  "text-blue-600"],
  ["Warehouses", s.wh_count,   "text-indigo-600"],
  ["Features",   s.feat_count, "text-purple-600"],
  ["Adoption",   s.pct + "%",  s.pct >= 70 ? "text-green-600" : s.pct >= 40 ? "text-yellow-600" : "text-red-600"],
].map(([label, val, cls]) =>
  `<div class="text-center">
     <div class="text-2xl font-bold ${cls}">${val}</div>
     <div class="text-xs text-gray-400 mt-0.5">${label}</div>
   </div>`
).join('<div class="w-px bg-gray-200 self-stretch mx-1"></div>');

/* ── Section toggle ───────────────────────────────────────────── */
function showSection(id) {
  ["overview","warehouse"].forEach(s => {
    document.getElementById("sec-"+s).classList.toggle("hidden", s !== id);
    const btn = document.getElementById("tab-"+s);
    btn.classList.toggle("tab-active", s === id);
    btn.classList.toggle("tab-idle",   s !== id);
  });
}

/* ── PCT cell class ───────────────────────────────────────────── */
function pctClass(pct) {
  if (pct === null)  return "pct-na";
  if (pct === 100)   return "pct-100";
  if (pct >= 80)     return "pct-high";
  if (pct >= 40)     return "pct-mid";
  if (pct > 0)       return "pct-low";
  return "pct-zero";
}

/* ── Floating tooltip portal ──────────────────────────────────── */
(function initTooltip() {
  const portal = document.getElementById("tip-portal");
  let active = null;

  document.addEventListener("mouseover", e => {
    const icon = e.target.closest(".tip-icon");
    if (!icon) return;
    const text = icon.dataset.tip;
    if (!text) return;
    active = icon;
    portal.textContent = text;
    portal.style.display = "block";
  });

  document.addEventListener("mousemove", e => {
    if (!active) return;
    const gap = 14;
    const pw  = portal.offsetWidth, ph = portal.offsetHeight;
    const vw  = window.innerWidth,   vh = window.innerHeight;
    let x = e.clientX + gap, y = e.clientY + gap;
    if (x + pw > vw - 8) x = e.clientX - pw - gap;
    if (y + ph > vh - 8) y = e.clientY - ph - gap;
    portal.style.left = x + "px";
    portal.style.top  = y + "px";
  });

  document.addEventListener("mouseout", e => {
    if (e.target.closest(".tip-icon")) {
      portal.style.display = "none";
      active = null;
    }
  });
})();

/* ── Tooltip cell helper ──────────────────────────────────────── */
function tipCell(label, description, extraClass = "") {
  const esc = description.replace(/"/g, "&quot;");
  return `<td class="sticky-col px-3 py-2 text-sm text-gray-800 whitespace-nowrap font-medium ${extraClass}">
    <span class="tip-wrap">
      ${label}
      <span class="tip-icon" data-tip="${esc}">?</span>
    </span>
  </td>`;
}

/* ── Overview table ───────────────────────────────────────────── */
(function buildOverview() {
  const { features, markets, feat_mkt, feat_total, domains, feat_desc } = DATA;

  let html = "<thead><tr>"
    + `<th class="sticky-col text-left px-3 py-2 text-xs font-semibold text-gray-600 whitespace-nowrap">Feature</th>`
    + `<th class="px-3 py-2 text-xs font-semibold text-gray-600 whitespace-nowrap">Domain</th>`
    + markets.map(m =>
        `<th class="px-3 py-2 text-xs font-semibold text-gray-600 text-center whitespace-nowrap">${m}</th>`
      ).join("")
    + `<th class="px-3 py-2 text-xs font-semibold text-gray-800 text-center whitespace-nowrap bg-gray-100 border-l-2 border-gray-300">Total</th>`
    + "</tr></thead><tbody>";

  let prevDomain = null;
  features.forEach((feat, i) => {
    const domain = domains[feat] || "";

    // Domain group separator row
    if (domain !== prevDomain) {
      html += `<tr><td colspan="${markets.length + 3}"
        style="background:#1e293b;color:#94a3b8;font-size:10px;font-weight:700;
               letter-spacing:.08em;padding:4px 12px;text-transform:uppercase;">
        ${domain || "Other"}
      </td></tr>`;
      prevDomain = domain;
    }

    const bg = i % 2 === 0 ? "bg-white" : "bg-gray-50/50";
    html += `<tr class="${bg}">`;
    html += tipCell(feat, feat_desc[feat] || "");
    html += `<td class="px-3 py-2 text-xs text-gray-400 whitespace-nowrap">${domain}</td>`;

    markets.forEach(mkt => {
      const d = feat_mkt[feat]?.[mkt];
      if (!d) {
        html += `<td class="px-3 py-2 pct-na text-center text-xs">—</td>`;
      } else {
        const cls = pctClass(d.pct);
        html += `<td class="px-3 py-2 ${cls} text-center text-xs font-bold" title="${d.yes}/${d.total} warehouses">${d.pct}%</td>`;
      }
    });

    // Total column
    const t = feat_total[feat];
    if (t) {
      const cls = pctClass(t.pct);
      html += `<td class="px-3 py-2 ${cls} text-center text-xs font-bold border-l-2 border-gray-300"
                   title="${t.yes}/${t.total} warehouses across all markets">${t.pct}%<br>
               <span style="font-weight:400;font-size:9px;opacity:.85">${t.yes}/${t.total}</span></td>`;
    } else {
      html += `<td class="px-3 py-2 pct-na text-center text-xs border-l-2 border-gray-300">—</td>`;
    }

    html += "</tr>";
  });
  html += "</tbody>";
  document.getElementById("overviewTable").innerHTML = html;
})();

/* ── Warehouse tables ─────────────────────────────────────────── */
(function buildWarehouse() {
  const { features, wh_by_mkt, cells, domains, feat_desc } = DATA;
  const markets = Object.keys(wh_by_mkt);
  const tabsEl  = document.getElementById("mktTabs");
  const tbls    = document.getElementById("wh-tables");

  markets.forEach((mkt, idx) => {
    // tab button
    const btn = document.createElement("button");
    btn.id        = `mkt-tab-${mkt}`;
    btn.textContent = mkt;
    btn.className = (idx === 0 ? "tab-active" : "tab-idle") + " px-4 py-2 rounded-lg text-sm font-medium";
    btn.onclick   = () => showMkt(mkt);
    tabsEl.appendChild(btn);

    // table wrapper
    const wrap = document.createElement("div");
    wrap.id = `mkt-${mkt}`;
    wrap.className = (idx === 0 ? "" : "hidden") + " scroll-x";

    const whs = wh_by_mkt[mkt];
    let html = "<table class='w-full text-xs'><thead><tr>"
      + `<th class="sticky-col px-3 py-2 text-left font-semibold text-gray-600">Warehouse</th>`
      + features.map(f => {
          const short = f.length > 18 ? f.slice(0,17)+"…" : f;
          const dom   = domains[f] || "";
          const desc  = (feat_desc[f] || "").replace(/"/g, "&quot;");
          return `<th class="px-2 py-2 font-semibold text-gray-600 text-center whitespace-nowrap">
            <span class="tip-wrap" style="justify-content:center">
              <span>${short}</span>
              <span class="tip-icon" data-tip="${desc}">?</span>
            </span>
            <br><span style="font-weight:400;font-size:9px;color:#94a3b8">${dom}</span>
          </th>`;
        }).join("")
      + "</tr></thead><tbody>";

    whs.forEach((wh, i) => {
      const bg = i % 2 === 0 ? "bg-white" : "bg-gray-50/50";
      html += `<tr class="${bg}"><td class="sticky-col px-3 py-2 font-medium text-gray-800 whitespace-nowrap">${wh}</td>`;
      features.forEach(feat => {
        const val = cells[`${mkt}|${wh}|${feat}`];
        if (val === undefined) {
          html += `<td class="cell-na px-2 py-2 text-center">—</td>`;
        } else if (val === "Yes") {
          html += `<td class="cell-yes px-2 py-2 text-center">✓</td>`;
        } else if (val === "No") {
          html += `<td class="cell-no px-2 py-2 text-center">✗</td>`;
        } else {
          html += `<td class="cell-err px-2 py-2 text-center">!</td>`;
        }
      });
      html += "</tr>";
    });
    html += "</tbody></table>";
    wrap.innerHTML = html;
    tbls.appendChild(wrap);
  });

  function showMkt(active) {
    markets.forEach(m => {
      document.getElementById(`mkt-${m}`).classList.toggle("hidden", m !== active);
      const btn = document.getElementById(`mkt-tab-${m}`);
      btn.classList.toggle("tab-active", m === active);
      btn.classList.toggle("tab-idle",   m !== active);
    });
  }
})();
</script>
</body>
</html>
"""


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate WMS Adoption HTML dashboard")
    parser.add_argument("--input",  default=str(DEFAULT_INPUT),  help="Path to wms_adoption_results.xlsx")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output HTML path (default: ../docs/index.html)")
    args = parser.parse_args()

    xlsx_path = Path(args.input)
    out_path  = Path(args.output)

    if not xlsx_path.exists():
        sys.exit(f"❌  Input file not found: {xlsx_path}")

    print(f"📂  Reading  {xlsx_path}")
    results = load_results(xlsx_path)
    if not results:
        sys.exit("❌  No rows found in Detail sheet.")

    print(f"✅  Loaded   {len(results)} rows")
    payload = build_payload(results)

    s = payload["stats"]
    print(f"    Markets: {s['mkt_count']}  |  WHs: {s['wh_count']}  |  Features: {s['feat_count']}")
    print(f"    Adoption: {s['adopted']}/{s['total']}  ({s['pct']}%)")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    html = HTML_TEMPLATE.replace("__DATA__", json.dumps(payload, ensure_ascii=False))
    out_path.write_text(html, encoding="utf-8")
    print(f"\n🌐  Dashboard → {out_path}")
    print(f"\n── GitHub Pages setup ────────────────────────────────────")
    print(f"   1. Commit & push the 'docs/' folder to your repo")
    print(f"   2. Repo Settings → Pages → Source: 'Deploy from branch'")
    print(f"   3. Branch: main  |  Folder: /docs  → Save")
    print(f"   Your dashboard will be live at:")
    print(f"   https://<your-org>.github.io/<repo-name>/")


if __name__ == "__main__":
    main()
