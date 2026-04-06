#!/usr/bin/env python3
"""
WMS Adoption Tracker — Engine
==============================
Write once. Never edit.

- Add a feature  → drop a new file in checkers/
- Change markets → edit features.py
- Change config  → edit features.py

Usage:
    python3 runner.py
"""

import asyncio
import datetime
import importlib.util
import pathlib
from pathlib import Path
from collections import defaultdict

from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from features import FEATURES, MARKETS

# ─── INFRASTRUCTURE CONSTANTS ─────────────────────────────────────────────────
# Only change these if Shopee changes its domain structure.

WMS_BASE_URLS = {
    "ID": "https://wms.ssc.shopee.co.id/v2",
    "MY": "https://wms.ssc.shopee.com.my/v2",
    "TH": "https://wms.ssc.shopee.co.th/v2",
    "PH": "https://wms.ssc.shopee.ph/v2",
    "SG": "https://wms.ssc.shopee.sg/v2",
    "VN": "https://wms.ssc.shopee.vn/v2",
    "TW": "https://wms.ssc.shopee.tw/v2",
    "BR": "https://wms.ssc.shopee.com.br/v2",
}

GUARDED_HOSTS = (
    {url.split("/")[2] for url in WMS_BASE_URLS.values()}
    | {"ops.ssc.shopeemobile.com"}
)

HERE         = Path(__file__).parent
PROFILE_DIR  = str(HERE / "wms_profile")
CHECKERS_DIR = HERE / "checkers"
OUTPUT_FILE  = str(HERE / "output" / "wms_adoption_results.xlsx")


# ─── SAFETY GUARD ─────────────────────────────────────────────────────────────
# Block all network mutations on WMS and admin portal.
# POSTs to read-like endpoints are allowed; POSTs to write-like paths are not.

_WRITE_VERBS     = {"create", "update", "delete", "remove", "save", "submit",
                    "edit", "modify", "add_", "batch_update", "upload"}
_ALLOWED_PATHS   = {"set_user_setting"}   # warehouse-switch API
_BLOCKED_METHODS = {"PUT", "DELETE", "PATCH"}


async def _guard(route, request):
    host = request.url.split("/")[2] if "//" in request.url else ""
    if host not in GUARDED_HOSTS:
        await route.continue_()
        return
    method = request.method.upper()
    path   = request.url.lower()
    is_allowed    = any(a in path for a in _ALLOWED_PATHS)
    is_write_post = method == "POST" and any(v in path for v in _WRITE_VERBS)
    if not is_allowed and (method in _BLOCKED_METHODS or is_write_post):
        print(f"\n  🛡️  BLOCKED {method} → {request.url[:80]}")
        await route.abort()
    else:
        await route.continue_()


# ─── CHECKER AUTO-DISCOVERY ───────────────────────────────────────────────────

def load_checkers() -> dict:
    """
    Import every *.py file in checkers/ (skip _*.py).
    Returns {FEATURE_NAME: module}.
    Each module must expose: FEATURE_NAME, CHECK_TYPE, SIGNAL,
    and async def check(page, warehouse, market, params).
    """
    checkers = {}
    for path in sorted(CHECKERS_DIR.glob("*.py")):
        if path.stem.startswith("_"):
            continue
        spec = importlib.util.spec_from_file_location(path.stem, path)
        mod  = importlib.util.module_from_spec(spec)
        try:
            spec.loader.exec_module(mod)
            for attr in ("FEATURE_NAME", "CHECK_TYPE", "SIGNAL", "check"):
                if not hasattr(mod, attr):
                    raise AttributeError(f"missing required attribute: {attr}")
            checkers[mod.FEATURE_NAME] = mod
            print(f"  ✓  {path.name:35s}  →  {mod.FEATURE_NAME}")
        except Exception as e:
            print(f"  ✗  {path.name}: {e}")
    return checkers


# ─── WAREHOUSE HELPERS ────────────────────────────────────────────────────────

async def _discover_warehouses(page, market: str, base_url: str) -> list:
    """Read all warehouse codes from the WMS sidebar dropdown."""
    await page.goto(base_url, wait_until="networkidle", timeout=60000)
    await page.wait_for_timeout(1000)

    opened = await page.evaluate("""() => {
        const el = Array.from(document.querySelectorAll('*')).find(e =>
            e.offsetParent && /^[A-Z]{2} - [A-Z0-9]+$/.test((e.innerText||'').trim())
        );
        if (!el) return false;
        el.click();
        return true;
    }""")
    if not opened:
        print(f"  ⚠️  Could not open warehouse dropdown for {market}")
        return []

    await page.wait_for_timeout(600)
    options = await page.evaluate("""() =>
        Array.from(document.querySelectorAll('li.ssc-option, li[class*="ssc-option"]'))
            .map(el => el.innerText.trim())
    """)
    await page.keyboard.press("Escape")

    prefix = f"{market} - "
    codes  = [o[len(prefix):].strip() for o in options if o.startswith(prefix)]
    print(f"  📦 Discovered {len(codes)} warehouses for {market}: {codes}")
    return codes


async def _navigate_and_switch_warehouse(page, url: str, market: str, whs: str):
    """Navigate to url, switch warehouse, then re-navigate to url."""
    await page.goto(url, wait_until="networkidle", timeout=60000)
    await page.wait_for_timeout(800)

    # If the goto redirected to a login page, wait for the user to log in
    if _is_login_page(page.url):
        await _ensure_logged_in(page, url, f"{market} WMS")

    # Open trigger
    trigger_ok = await page.evaluate("""() => {
        const el = Array.from(document.querySelectorAll('*')).find(e =>
            e.offsetParent && /^[A-Z]{2} - [A-Z0-9]+$/.test((e.innerText||'').trim())
        );
        if (!el) return false;
        el.click();
        return true;
    }""")
    if not trigger_ok:
        raise RuntimeError("warehouse trigger not visible — session may have expired")

    await page.wait_for_timeout(600)

    # Select option
    target = f"{market} - {whs}"
    picked = await page.evaluate("""(target) => {
        const opts = Array.from(
            document.querySelectorAll('li.ssc-option, li[class*="ssc-option"]')
        );
        const el = opts.find(e => (e.innerText||'').trim() === target);
        if (!el) {
            return {ok: false,
                    available: opts.map(e => e.innerText.trim()).slice(0, 8)};
        }
        el.click();
        return {ok: true};
    }""", target)

    if not picked.get("ok"):
        raise RuntimeError(
            f"option '{target}' not found. Available: {picked.get('available')}"
        )

    await page.wait_for_load_state("networkidle", timeout=20000)
    await page.wait_for_timeout(1000)

    # Re-navigate: warehouse switch may reload the page
    await page.goto(url, wait_until="networkidle", timeout=60000)
    await page.wait_for_timeout(500)


# ─── LOGIN HELPERS ────────────────────────────────────────────────────────────

def _is_login_page(url: str) -> bool:
    u = url.lower()
    return "login" in u or "sso" in u or "accounts.google" in u


async def _wait_for_wms_app(page, timeout_s: int = 120):
    """
    Poll until a WMS UI element is visible (warehouse selector or sidebar).
    This is more reliable than checking the URL because the post-OAuth redirect
    can land on the WMS page without the URL passing _is_login_page().
    """
    import time
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        try:
            found = await page.evaluate("""() =>
                document.querySelector('.ssc-select, [class*="ssc-select"]') !== null
                || document.querySelector('.ant-layout-sider') !== null
            """)
            if found:
                return
        except Exception:
            pass
        await asyncio.sleep(1)
    raise TimeoutError("WMS app did not load — login may not have completed")


async def _ensure_logged_in(page, url: str, label: str = ""):
    """Navigate to url; if redirected to login, wait for the user to log in."""
    await page.goto(url, wait_until="networkidle", timeout=60000)
    if not _is_login_page(page.url):
        return
    print(f"\n{'='*60}")
    print(f"⚠️  Login required — {label or url}")
    print("  Complete Google SSO in the browser window.")
    print("  (The script will continue automatically once WMS loads.)")
    print(f"{'='*60}")
    await _wait_for_wms_app(page)
    print("  ✓ Login detected — continuing.")


# ─── MAIN RUN LOOP ────────────────────────────────────────────────────────────

async def run():
    Path(OUTPUT_FILE).parent.mkdir(parents=True, exist_ok=True)
    checked_at = datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    results    = []

    print("\n── Loading checkers " + "─" * 36)
    checkers = load_checkers()
    if not checkers:
        print("No checkers found in checkers/. Exiting.")
        return
    print(f"── {len(checkers)} checker(s) loaded " + "─" * 29)

    async with async_playwright() as p:
        is_first_run = not Path(PROFILE_DIR).exists()

        context = await p.chromium.launch_persistent_context(
            user_data_dir=PROFILE_DIR,
            channel="chrome",
            headless=False,
            args=["--disable-blink-features=AutomationControlled"],
        )
        page = await context.new_page()
        await page.route("**/*", _guard)

        # First-run: visit every required WMS market + admin portal
        # so the user can log in to all of them before the run starts.
        needed_markets = {m for f in FEATURES for m in f["markets"]}
        needs_admin    = any(
            getattr(checkers.get(f["feature"]), "CHECK_TYPE", "") == "admin_portal"
            for f in FEATURES
        )

        if is_first_run:
            print("\n" + "=" * 60)
            print("FIRST RUN — log in to every required portal")
            print("=" * 60)

        for market in sorted(needed_markets):
            if market not in WMS_BASE_URLS:
                continue
            url = WMS_BASE_URLS[market]
            await page.goto(url, wait_until="networkidle", timeout=60000)
            if _is_login_page(page.url):
                print(f"\n  🔐 {market} WMS needs login ({url})")
                await _wait_for_wms_app(page)
                print(f"  ✓  {market} WMS — logged in")
            else:
                print(f"  ✓  {market} WMS — session OK")

        if needs_admin:
            admin_url = "https://ops.ssc.shopeemobile.com"
            await page.goto(admin_url, wait_until="networkidle", timeout=60000)
            if _is_login_page(page.url):
                print(f"\n  🔐 Admin portal needs login ({admin_url})")
                print("  Complete login in browser, then the script continues automatically.")
                # Admin portal may not have .ssc-select; wait for URL to leave login
                for _ in range(120):
                    if not _is_login_page(page.url):
                        break
                    await asyncio.sleep(1)
                print("  ✓  Admin portal — logged in")
            else:
                print("  ✓  Admin portal — session OK")

        print(f"\nProfile: {PROFILE_DIR}\n")

        # Auto-discover warehouses for markets without a hardcoded list
        needed_markets = {m for f in FEATURES for m in f["markets"]}
        for market in sorted(needed_markets):
            if MARKETS.get(market):
                continue
            if market not in WMS_BASE_URLS:
                print(f"⚠️  No WMS URL for {market} — skipping discovery")
                continue
            print(f"Discovering warehouses for {market}...")
            await _ensure_logged_in(page, WMS_BASE_URLS[market], market)
            MARKETS[market] = await _discover_warehouses(
                page, market, WMS_BASE_URLS[market]
            )

        # ── Feature loop ──────────────────────────────────────────────────────
        for feature_cfg in FEATURES:
            feature_name = feature_cfg["feature"]
            checker      = checkers.get(feature_name)

            if not checker:
                print(f"\n⚠️  No checker for '{feature_name}' — skipping")
                continue

            check_type = checker.CHECK_TYPE
            signal     = checker.SIGNAL
            params     = feature_cfg.get("params", {})

            for market in feature_cfg["markets"]:
                warehouses = MARKETS.get(market, [])
                if not warehouses:
                    print(f"\n⚠️  No warehouses for {market} — skipping '{feature_name}'")
                    continue

                print(f"\n{'─'*58}")
                print(f"  Feature : {feature_name}")
                print(f"  Market  : {market}  ({len(warehouses)} warehouses)  [{check_type}]")
                print(f"{'─'*58}")

                # Pre-flight: ensure the market portal is accessible before looping
                if check_type == "wms_frontend" and market in WMS_BASE_URLS:
                    base = WMS_BASE_URLS[market]
                    await page.goto(base, wait_until="networkidle", timeout=60000)
                    if _is_login_page(page.url):
                        print(f"\n  🔐 {market} WMS session expired — please log in.")
                        await _wait_for_wms_app(page)

                for i, whs in enumerate(warehouses):
                    try:
                        print(f"  [{i+1:2d}/{len(warehouses)}] {whs:10s}", end=" ", flush=True)

                        # Runner owns navigation for wms_frontend only
                        if check_type == "wms_frontend":
                            url = WMS_BASE_URLS[market] + params["url_path"]
                            await _navigate_and_switch_warehouse(page, url, market, whs)

                        sig_val, adopted = await checker.check(page, whs, market, params)
                        mark = "✅" if adopted else ("⚠️" if adopted is None else "❌")
                        print(f"→ {sig_val} {mark}")

                        results.append({
                            "market": market, "warehouse": whs,
                            "feature": feature_name, "signal": signal,
                            "signal_value": sig_val, "adopted": adopted,
                            "checked_at": checked_at,
                        })

                    except Exception as e:
                        print(f"→ ERROR: {e} ❌")
                        results.append({
                            "market": market, "warehouse": whs,
                            "feature": feature_name, "signal": signal,
                            "signal_value": f"error: {e}", "adopted": None,
                            "checked_at": checked_at,
                        })

        await context.close()

    _save_excel(results, checked_at)
    total   = len(results)
    adopted = sum(1 for r in results if r["adopted"] is True)
    valid   = sum(1 for r in results if r["adopted"] is not None)
    print(f"\n✅  Results → {OUTPUT_FILE}")
    print(f"    Adopted: {adopted}/{valid}  ({int(adopted/valid*100) if valid else 0}%)  "
          f"| Errors: {total - valid}")


# ─── EXCEL OUTPUT ─────────────────────────────────────────────────────────────

def _save_excel(results: list, checked_at: str):
    wb = Workbook()

    # ── Detail sheet ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Detail"
    headers  = ["Market", "Warehouse", "Feature", "Signal", "Signal Value", "Adopted", "Checked At"]

    green = PatternFill("solid", fgColor="C6EFCE")
    red   = PatternFill("solid", fgColor="FFC7CE")
    warn  = PatternFill("solid", fgColor="FFEB9C")
    grey  = PatternFill("solid", fgColor="D9D9D9")
    gf    = Font(name="Arial", color="276221")
    rf    = Font(name="Arial", color="9C0006")
    wf    = Font(name="Arial", color="9C6500")
    bold  = Font(name="Arial", bold=True)
    base  = Font(name="Arial")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = bold; cell.fill = grey
        cell.alignment = Alignment(horizontal="center")

    for ri, r in enumerate(results, 2):
        adopted = r["adopted"]
        fill, font = (green, gf) if adopted else ((red, rf) if adopted is False else (warn, wf))
        row_vals = [
            r["market"], r["warehouse"], r["feature"], r["signal"],
            r["signal_value"],
            "Yes" if adopted is True else ("No" if adopted is False else "Error"),
            r["checked_at"],
        ]
        for ci, v in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = font; cell.fill = fill
            cell.alignment = Alignment(horizontal="left" if ci in (3, 4, 5) else "center")

    tot  = len(results)
    yes  = sum(1 for r in results if r["adopted"] is True)
    sr   = tot + 2
    ws.cell(row=sr, column=1, value="TOTAL").font = bold
    ws.cell(row=sr, column=6,
            value=f"{yes}/{tot} ({int(yes/tot*100) if tot else 0}%)").font = bold

    for i, w in enumerate([8, 12, 32, 46, 28, 10, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    hdr = ["Feature", "Market", "Total", "Adopted", "Not Adopted", "Errors", "Adoption %"]
    for c, h in enumerate(hdr, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = bold; cell.fill = grey
        cell.alignment = Alignment(horizontal="center")

    by_fm = defaultdict(list)
    for r in results:
        by_fm[(r["feature"], r["market"])].append(r)

    for ri, ((feat, mkt), rows) in enumerate(sorted(by_fm.items()), 2):
        t   = len(rows)
        yes = sum(1 for r in rows if r["adopted"] is True)
        no  = sum(1 for r in rows if r["adopted"] is False)
        err = sum(1 for r in rows if r["adopted"] is None)
        pct = f"{int(yes/t*100)}%" if t else "0%"
        for c, v in enumerate([feat, mkt, t, yes, no, err, pct], 1):
            cell = ws2.cell(row=ri, column=c, value=v)
            cell.font = base
            cell.alignment = Alignment(horizontal="left" if c == 1 else "center")

    for i, w in enumerate([36, 8, 8, 10, 14, 8, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUTPUT_FILE)


# ─── ENTRY POINT ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    asyncio.run(run())
