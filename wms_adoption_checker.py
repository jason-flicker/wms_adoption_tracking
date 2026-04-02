#!/usr/bin/env python3
"""
WMS Feature Adoption Checker
Scalable: configure FEATURES and MARKETS below, run once.

Check types
-----------
wms_frontend   — per-warehouse check on wms.ssc.shopee.<tld>
admin_portal   — single-page check on ops.ssc.shopeemobile.com
                 (one page shows all warehouses; region config takes precedence)

First run:  opens a real Chrome window (dedicated profile), log in via
            Google SSO on both WMS and admin portal, then press Enter.
Later runs: profile is reused automatically (no login needed).

Usage:
    python3 wms_adoption_checker.py

Output:
    wms_adoption_results.xlsx
"""

import re
import asyncio
import datetime
from pathlib import Path
from collections import defaultdict
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─── CONFIG ──────────────────────────────────────────────────────────────────

WMS_BASE_URLS = {
    "ID": "https://wms.ssc.shopee.co.id/v2",
    "MY": "https://wms.ssc.shopee.com.my/v2",
    "TH": "https://wms.ssc.shopee.co.th/v2",
    "PH": "https://wms.ssc.shopee.ph/v2",
    "SG": "https://wms.ssc.shopee.sg/v2",
    "VN": "https://wms.ssc.shopee.vn/v2",
    "TW": "https://wms.ssc.shopee.tw/v2",
    "BR": "https://wms.ssc.shopee.com.br/v2",
}

ADMIN_PORTAL_BASE = "https://ops.ssc.shopeemobile.com"

# All WMS + ops portal hosts — safety guard is applied to all of them
GUARDED_HOSTS = (
    {url.split("/")[2] for url in WMS_BASE_URLS.values()} | {"ops.ssc.shopeemobile.com"}
)

FEATURES = [
    # ── wms_frontend ─────────────────────────────────────────────────────────
    {
        "feature": "MTO Picking Task Generation",
        "check_type": "wms_frontend",
        "url_path": "/rulecenter/pickingrule/mtoPickingRule",
        "checker": "text_label_yes_no",
        "signal": "Split MTO into multiple picking tasks",
        "markets": ["PH", "SG", "MY"],
    },
    {
        "feature": "Dynamic Wave Rule Setting",
        "check_type": "wms_frontend",
        "url_path": "/rulecenter/waverule/dynamicWaveRule",
        "checker": "table_has_active_row",
        "signal": "At least 1 active rule in Dynamic Wave Rule table",
        "markets": ["SG", "MY"],
    },
    {
        "feature": "Picking While Sorting",
        "check_type": "wms_frontend",
        "url_path": "/rulecenter/skillManagementRule/operatorSkill/salesOutbound/picking",
        "checker": "filter_count_gte",
        "signal": ">=5 users with Picking Method = Sorting While Picking",
        "filter_label": "Picking Method",
        "filter_value": "Sorting While Picking",
        "min_count": 5,
        "markets": ["SG", "MY"],
    },
    {
        "feature": "Dynamic Replenishment",
        "check_type": "wms_frontend",
        "url_path": "/inventorymanage/racktransfer/order",
        "checker": "filter_has_recent",
        "signal": "Orders from Replenishment Demand Pool in last 7 days",
        "filter_label": "Source from",
        "filter_value": "Replenishment Demand Pool",
        "days": 7,
        "markets": ["SG", "MY"],
    },
    {
        "feature": "MTO Exclusion of Non-Sellable Stock",
        "check_type": "wms_frontend",
        "url_path": "/rulecenter/allocateRule/mt",
        "checker": "has_toggled_on_rule",
        "signal": "At least 1 Allocate Exclusion Rule toggled on",
        "markets": ["SG", "MY"],
    },
    # ── admin_portal ─────────────────────────────────────────────────────────
    {
        "feature": "Inbound Boxing",
        "check_type": "admin_portal",
        "checker": "admin_portal_config",
        "signal": "Non_QC_Item_Putaway_Directly = 1",
        "checks": [
            {
                "url": "https://ops.ssc.shopeemobile.com/wms/configurationmanagement/configuration/view?conf_key=Non_QC_Item_Putaway_Directly",
                "key": "Non_QC_Item_Putaway_Directly",
                "target_value": "1",
            }
        ],
        "markets": ["SG", "MY"],
    },
    {
        "feature": "Dynamic Wave Toggle",
        "check_type": "admin_portal",
        "checker": "admin_portal_config",
        "signal": "wave_dynamic_wave_task_enable=1 AND wave_algorithm_switch=1 AND pre_allocate_zone_inventory=1",
        "checks": [
            {
                "url": "https://ops.ssc.shopeemobile.com/wms/configurationmanagement/configuration/view?conf_key=wave_dynamic_wave_task_enable",
                "key": "wave_dynamic_wave_task_enable",
                "target_value": "1",
            },
            {
                "url": "https://ops.ssc.shopeemobile.com/wms/configurationmanagement/configuration/view?conf_key=wave_algorithm_switch",
                "key": "wave_algorithm_switch",
                "target_value": "1",
            },
            {
                "url": "https://ops.ssc.shopeemobile.com/wms/configurationmanagement/configuration/view?conf_key=pre_allocate_zone_inventory",
                "key": "pre_allocate_zone_inventory",
                "target_value": "1",
            },
        ],
        "markets": ["SG", "MY"],
    },
]

MARKETS = {
    "PH": [
        "PHA", "PHB", "PHD", "PHE", "PHG", "PHIXC", "PHIXP",
        "PHJ", "PHK", "PHL", "PHM", "PHN", "PHO", "PHP",
        "PHR", "PHS", "PHT", "PHTSDS", "PHU", "PHV", "PHX", "PHY",
    ],
    "SG": [],   # auto-discovered from WMS dropdown on first access
    "MY": [],   # auto-discovered from WMS dropdown on first access
}

PROFILE_DIR = str(Path(__file__).parent / "wms_profile")
OUTPUT_FILE  = str(Path(__file__).parent / "wms_adoption_results.xlsx")

# ─── SAFETY GUARD ────────────────────────────────────────────────────────────

WRITE_VERBS    = {"create", "update", "delete", "remove", "save", "submit",
                  "edit", "modify", "add_", "batch_update", "upload"}
ALLOWED_PATHS  = {"set_user_setting"}
BLOCKED_METHODS = {"PUT", "DELETE", "PATCH"}


def make_guard(page):
    """Attach a network-level read-only guard to *page*."""
    async def block_mutations(route, request):
        host = request.url.split("/")[2] if "//" in request.url else ""
        if host not in GUARDED_HOSTS:
            await route.continue_()
            return
        method = request.method.upper()
        path   = request.url.lower()
        is_allowed    = any(a in path for a in ALLOWED_PATHS)
        is_write_post = method == "POST" and any(v in path for v in WRITE_VERBS)
        if not is_allowed and (method in BLOCKED_METHODS or is_write_post):
            print(f"\n  🛡️  BLOCKED {method} → {request.url[:80]}")
            await route.abort()
        else:
            await route.continue_()
    return page.route("**/*", block_mutations)


# ─── WAREHOUSE DISCOVERY ─────────────────────────────────────────────────────

async def discover_warehouses(page, market: str, base_url: str) -> list:
    """Read all warehouse codes for *market* from the WMS dropdown."""
    await page.goto(base_url, wait_until="networkidle", timeout=30000)
    await page.wait_for_timeout(1000)

    click_result = await page.evaluate("""() => {
        const el = Array.from(document.querySelectorAll('*')).find(e => {
            if (!e.offsetParent) return false;
            return /^[A-Z]{2} - [A-Z0-9]+$/.test((e.innerText || '').trim());
        });
        if (!el) return {ok: false};
        el.click();
        return {ok: true};
    }""")
    if not click_result.get("ok"):
        print(f"  ⚠️  Could not open dropdown for {market} — no warehouses discovered")
        return []

    await page.wait_for_timeout(600)
    options = await page.evaluate("""() =>
        Array.from(document.querySelectorAll('li.ssc-option, li[class*="ssc-option"]'))
            .map(el => el.innerText.trim())
    """)
    await page.keyboard.press("Escape")

    prefix = f"{market} - "
    codes = [o[len(prefix):].strip() for o in options if o.startswith(prefix)]
    print(f"  📦 Discovered {len(codes)} warehouses for {market}: {codes}")
    return codes


# ─── WMS FRONTEND CHECKERS ───────────────────────────────────────────────────

async def check_text_label_yes_no(page, feature: dict) -> tuple:
    """Find a label text on the page then read the adjacent Yes/No value."""
    signal = feature["signal"]
    try:
        await page.wait_for_function(
            f"document.body.innerText.includes({repr(signal)})",
            timeout=15000,
        )
    except Exception:
        preview = (await page.evaluate("document.body.innerText"))[:200]
        return "element_not_found", None

    text  = await page.evaluate("document.body.innerText")
    match = re.search(
        re.escape(signal) + r"[\s\n]+(Yes|No|YES|NO|Enabled|Disabled|On|Off)",
        text, re.IGNORECASE,
    )
    if not match:
        return "unknown", None
    val = match.group(1).capitalize()
    return val, val.lower() in ("yes", "enabled", "on")


async def check_table_has_active_row(page, feature: dict) -> tuple:
    """Check whether any table row contains an 'Active' status cell."""
    try:
        await page.wait_for_load_state("networkidle", timeout=15000)
        result = await page.evaluate("""() => {
            const cells = Array.from(document.querySelectorAll('td, .ant-table-cell, [class*="table-cell"]'));
            const active = cells.find(el => /^active$/i.test((el.innerText || '').trim()));
            return {found: !!active, total_cells: cells.length};
        }""")
        if result["total_cells"] == 0:
            return "page_no_table", None
        if result["found"]:
            return "active_rule_found", True
        return "no_active_rule", False
    except Exception as e:
        return f"error: {e}", None


async def check_has_toggled_on_rule(page, feature: dict) -> tuple:
    """Check whether any toggle switch is in the ON state."""
    try:
        await page.wait_for_load_state("networkidle", timeout=15000)
        result = await page.evaluate("""() => {
            // Ant Design: .ant-switch-checked = ON
            const on = document.querySelectorAll(
                '.ant-switch-checked, input[type="checkbox"]:checked'
            );
            return {on_count: on.length};
        }""")
        count = result["on_count"]
        if count > 0:
            return f"{count}_rules_toggled_on", True
        return "no_rules_toggled_on", False
    except Exception as e:
        return f"error: {e}", None


async def check_filter_count_gte(page, feature: dict) -> tuple:
    """
    Apply a dropdown filter then count matching rows.
    NOTE: Filter interaction depends on the specific page's UI.
          Extend this function once the page structure is confirmed.
    """
    # TODO: click filter dropdown, select feature["filter_value"], count rows
    return "filter_check_not_implemented", None


async def check_filter_has_recent(page, feature: dict) -> tuple:
    """
    Apply a filter and check for rows with Create Time in the last N days.
    NOTE: Date-range filter interaction TBD.
    """
    # TODO: apply filter, read dates, compare against feature["days"]
    return "filter_check_not_implemented", None


# Dispatcher — maps checker name → function
CHECKER_DISPATCH = {
    "text_label_yes_no":   check_text_label_yes_no,
    "table_has_active_row": check_table_has_active_row,
    "has_toggled_on_rule": check_has_toggled_on_rule,
    "filter_count_gte":    check_filter_count_gte,
    "filter_has_recent":   check_filter_has_recent,
}


# ─── ADMIN PORTAL CHECKER ────────────────────────────────────────────────────

async def check_admin_portal_config(page, feature: dict, market: str, warehouses: list) -> dict:
    """
    Evaluate one or more admin-portal config pages for a given market.

    Adoption logic (per KB):
      1. If region (market code) is configured with target value → all warehouses adopted.
      2. Else check each warehouse individually.
      3. All checks in feature["checks"] must pass (logical AND) for adopted=True.

    Returns: {warehouse_code: (signal_value, adopted)}
    """
    # Accumulate per-warehouse results across all sub-checks
    # adopted[whs] = True only if every check passes
    per_whs_results = {whs: [] for whs in warehouses}

    for check in feature["checks"]:
        url          = check["url"]
        key          = check["key"]
        target_value = check["target_value"]

        await page.goto(url, wait_until="networkidle", timeout=30000)
        await page.wait_for_timeout(1000)

        if "login" in page.url.lower() or "sso" in page.url.lower():
            print(f"\n  ⚠️  Admin portal login required. Please log in, then press Enter.")
            input()
            await page.goto(url, wait_until="networkidle", timeout=30000)

        text = await page.evaluate("document.body.innerText")

        # Region-level check: does the market code appear near the target value?
        region_adopted = bool(
            re.search(rf"(?<!\w){re.escape(market)}(?!\w).*?\b{re.escape(target_value)}\b",
                      text, re.IGNORECASE)
        )

        for whs in warehouses:
            if region_adopted:
                per_whs_results[whs].append((f"region_{market}={target_value}", True))
            else:
                whs_adopted = bool(
                    re.search(rf"(?<!\w){re.escape(whs)}(?!\w).*?\b{re.escape(target_value)}\b",
                              text, re.IGNORECASE)
                )
                per_whs_results[whs].append(
                    (f"{key}={target_value}" if whs_adopted else f"{key}!='{target_value}'",
                     whs_adopted)
                )

    # AND across all sub-checks
    final = {}
    for whs, check_results in per_whs_results.items():
        all_pass = all(r[1] for r in check_results)
        summary  = " & ".join(r[0] for r in check_results)
        final[whs] = (summary, all_pass)
    return final


# ─── LOGIN HELPERS ───────────────────────────────────────────────────────────

async def ensure_logged_in(page, url: str, prompt: str):
    await page.goto(url, wait_until="networkidle", timeout=30000)
    if "login" in page.url.lower() or "sso" in page.url.lower():
        print(f"\n⚠️  {prompt}")
        print("Press Enter here once you are fully logged in.")
        input()
        await page.goto(url, wait_until="networkidle", timeout=30000)


async def handle_mid_run_login(page, url: str):
    """Called when a navigation redirects to login mid-run."""
    if "login" in page.url.lower() or "sso" in page.url.lower():
        print("\n⚠️  Session expired — please log in again, then press Enter.")
        input()
        await page.goto(url, wait_until="networkidle", timeout=30000)


# ─── MAIN RUNNER ─────────────────────────────────────────────────────────────

async def run():
    checked_at = datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    results    = []

    async with async_playwright() as p:
        is_first_run = not Path(PROFILE_DIR).exists()

        context = await p.chromium.launch_persistent_context(
            user_data_dir=PROFILE_DIR,
            channel="chrome",
            headless=False,
            args=["--disable-blink-features=AutomationControlled"],
        )

        page = await context.new_page()
        await make_guard(page)

        if is_first_run:
            print("=" * 60)
            print("FIRST RUN — manual login required")
            print("=" * 60)
            print("\n1. Log in to WMS (PH as default):")
            await page.goto(WMS_BASE_URLS["PH"], wait_until="networkidle", timeout=30000)
            print("   Complete Google SSO login in the browser.")
            print("\n2. Then navigate to the admin portal and log in:")
            print(f"   {ADMIN_PORTAL_BASE}")
            print("\nPress Enter here once you are logged into BOTH portals.")
            input()
            print("Profile saved. Subsequent runs won't need login.\n")
        else:
            print(f"Reusing saved profile from {PROFILE_DIR}\n")

        # ── collect unique markets needed across all features ──────────────
        needed_markets = set()
        for feat in FEATURES:
            needed_markets.update(feat["markets"])

        # ── auto-discover warehouses for markets with empty lists ──────────
        for market in needed_markets:
            if market in MARKETS and MARKETS[market]:
                continue  # already have the list
            if market not in WMS_BASE_URLS:
                print(f"⚠️  No WMS URL configured for market {market} — skipping")
                continue
            print(f"Discovering warehouses for {market}...")
            await ensure_logged_in(page, WMS_BASE_URLS[market],
                                   f"Login required for {market} WMS")
            codes = await discover_warehouses(page, market, WMS_BASE_URLS[market])
            MARKETS[market] = codes

        # ── run each feature ───────────────────────────────────────────────
        for feature in FEATURES:
            feature_name = feature["feature"]
            check_type   = feature["check_type"]
            signal       = feature["signal"]

            for market in feature["markets"]:
                warehouses = MARKETS.get(market, [])
                if not warehouses:
                    print(f"\n⚠️  No warehouses for {market} — skipping {feature_name}")
                    continue

                print(f"\n{'─'*54}")
                print(f"Feature : {feature_name}")
                print(f"Market  : {market}  |  Check: {check_type}")
                print(f"{'─'*54}")

                # ── ADMIN PORTAL ──────────────────────────────────────────
                if check_type == "admin_portal":
                    try:
                        whs_results = await check_admin_portal_config(
                            page, feature, market, warehouses
                        )
                        for whs, (sig_val, adopted) in whs_results.items():
                            mark = "✅" if adopted else "❌"
                            print(f"  {whs:12s} → {sig_val} {mark}")
                            results.append({
                                "market": market, "warehouse": whs,
                                "feature": feature_name, "signal": signal,
                                "signal_value": sig_val, "adopted": adopted,
                                "checked_at": checked_at,
                            })
                    except Exception as e:
                        print(f"  ❌ admin portal error: {e}")
                        for whs in warehouses:
                            results.append({
                                "market": market, "warehouse": whs,
                                "feature": feature_name, "signal": signal,
                                "signal_value": f"error: {e}", "adopted": None,
                                "checked_at": checked_at,
                            })

                # ── WMS FRONTEND ──────────────────────────────────────────
                elif check_type == "wms_frontend":
                    base_url  = WMS_BASE_URLS.get(market, "")
                    url       = base_url + feature["url_path"]
                    checker_fn = CHECKER_DISPATCH.get(feature["checker"])

                    for i, whs in enumerate(warehouses):
                        try:
                            print(f"  [{i+1}/{len(warehouses)}] {whs}...",
                                  end=" ", flush=True)

                            await page.goto(url, wait_until="networkidle", timeout=30000)
                            await page.wait_for_timeout(1000)
                            await handle_mid_run_login(page, url)

                            await select_warehouse_from_dropdown(page, market, whs)

                            # Re-navigate after warehouse switch
                            await page.goto(url, wait_until="networkidle", timeout=30000)
                            await page.wait_for_timeout(500)

                            sig_val, adopted = await checker_fn(page, feature)
                            mark = "✅" if adopted else ("⚠️" if adopted is None else "❌")
                            print(f"→ {sig_val} {mark}")

                            results.append({
                                "market": market, "warehouse": whs,
                                "feature": feature_name, "signal": signal,
                                "signal_value": sig_val, "adopted": adopted,
                                "checked_at": checked_at,
                            })

                        except Exception as e:
                            print(f"→ ERROR: {e} ❌")
                            results.append({
                                "market": market, "warehouse": whs,
                                "feature": feature_name, "signal": signal,
                                "signal_value": f"error: {e}", "adopted": None,
                                "checked_at": checked_at,
                            })

        await context.close()

    save_excel(results, checked_at)
    print(f"\n✅ Done! Results saved to: {OUTPUT_FILE}")
    adopted = sum(1 for r in results if r["adopted"] is True)
    total   = len([r for r in results if r["adopted"] is not None])
    print(f"   Adopted: {adopted} / {total}  (errors excluded)")


# ─── WAREHOUSE DROPDOWN INTERACTION ──────────────────────────────────────────

async def select_warehouse_from_dropdown(page, market: str, warehouse_code: str):
    """Open the warehouse selector and pick *warehouse_code* for *market*."""

    # Step 1 — click the visible trigger (shows current "XX - CODE" selection)
    click_result = await page.evaluate("""() => {
        const el = Array.from(document.querySelectorAll('*')).find(e => {
            if (!e.offsetParent) return false;
            return /^[A-Z]{2} - [A-Z0-9]+$/.test((e.innerText || '').trim());
        });
        if (!el) return {ok: false, msg: 'no visible trigger found'};
        el.click();
        return {ok: true, clicked: el.tagName + '.' + el.className,
                text: el.innerText.trim()};
    }""")

    if not click_result.get("ok"):
        raise Exception(f"Could not find warehouse trigger: {click_result.get('msg')}")

    await page.wait_for_timeout(600)

    # Step 2 — click the target option by its full label "XX - CODE"
    target_text = f"{market} - {warehouse_code}"
    option_result = await page.evaluate("""(target) => {
        const opts = Array.from(
            document.querySelectorAll('li.ssc-option, li[class*="ssc-option"]')
        );
        const el = opts.find(e => (e.innerText || '').trim() === target);
        if (!el) {
            return {ok: false, msg: 'option not found',
                    available: opts.map(e => e.innerText.trim()).slice(0, 10)};
        }
        el.click();
        return {ok: true, clicked: el.tagName + '.' + el.className};
    }""", target_text)

    if not option_result.get("ok"):
        raise Exception(
            f"Option '{target_text}' not found. "
            f"Available: {option_result.get('available')}"
        )

    await page.wait_for_load_state("networkidle", timeout=20000)
    await page.wait_for_timeout(1000)


# ─── EXCEL OUTPUT ─────────────────────────────────────────────────────────────

def save_excel(results: list, checked_at: str):
    wb  = Workbook()
    ws  = wb.active
    ws.title = "WMS Adoption"

    headers    = ["Market", "Warehouse", "Feature", "Signal", "Signal Value", "Adopted", "Checked At"]
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill   = PatternFill("solid", fgColor="FFC7CE")
    warn_fill  = PatternFill("solid", fgColor="FFEB9C")
    grey_fill  = PatternFill("solid", fgColor="D9D9D9")
    green_font = Font(name="Arial", color="276221")
    red_font   = Font(name="Arial", color="9C0006")
    warn_font  = Font(name="Arial", color="9C6500")
    bold       = Font(name="Arial", bold=True)
    arial      = Font(name="Arial")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = bold
        cell.fill      = grey_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(results, 2):
        adopted = r["adopted"]
        if adopted is True:
            fill, font = green_fill, green_font
        elif adopted is False:
            fill, font = red_fill, red_font
        else:
            fill, font = warn_fill, warn_font

        values = [
            r["market"], r["warehouse"], r["feature"],
            r["signal"], r["signal_value"],
            "Yes" if adopted is True else ("No" if adopted is False else "Error"),
            r["checked_at"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = font
            cell.fill      = fill
            cell.alignment = Alignment(
                horizontal="left" if col_idx in (3, 4, 5) else "center"
            )

    total   = len(results)
    adopted = sum(1 for r in results if r["adopted"] is True)
    sr      = total + 2
    ws.cell(row=sr, column=1, value="TOTAL").font = bold
    ws.cell(row=sr, column=6,
            value=f"{adopted}/{total} ({int(adopted/total*100) if total else 0}%)").font = bold

    widths = [8, 10, 30, 45, 28, 10, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ── Summary by feature ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary by Feature")
    hdr = ["Feature", "Market", "Total", "Adopted", "Not Adopted", "Error", "Adoption %"]
    for c, h in enumerate(hdr, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = bold; cell.fill = grey_fill
        cell.alignment = Alignment(horizontal="center")

    by_feat_market = defaultdict(list)
    for r in results:
        by_feat_market[(r["feature"], r["market"])].append(r)

    for row_i, ((feat, mkt), rows) in enumerate(sorted(by_feat_market.items()), 2):
        tot = len(rows)
        yes = sum(1 for r in rows if r["adopted"] is True)
        no  = sum(1 for r in rows if r["adopted"] is False)
        err = sum(1 for r in rows if r["adopted"] is None)
        pct = f"{int(yes/tot*100)}%" if tot else "0%"
        for c, v in enumerate([feat, mkt, tot, yes, no, err, pct], 1):
            cell = ws2.cell(row=row_i, column=c, value=v)
            cell.font = arial
            cell.alignment = Alignment(horizontal="left" if c == 1 else "center")

    for i, w in enumerate([35, 8, 8, 10, 14, 8, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUTPUT_FILE)


# ─── ENTRY POINT ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    asyncio.run(run())
